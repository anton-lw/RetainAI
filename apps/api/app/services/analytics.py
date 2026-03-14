"""Operational analytics and queue-building logic.

This service is the bridge between stored data and the day-to-day user-facing
workspace. It builds beneficiary risk cases, retention summaries, intervention
effectiveness views, and export-ready queue slices. If a maintainer wants to
understand why a case appears in the queue with a given rank, this is one of
the first files to inspect.
"""

from __future__ import annotations

import csv
import io
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from math import ceil
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app import schemas
from app.core.time import utc_isoformat
from app.models import Beneficiary, Intervention, Program, ProgramOperationalSetting
from app.services.labeling import (
    build_operational_settings_profile,
    default_operational_profile,
    project_tracing_protocol,
    tracing_recommended_action,
)
from app.services.modeling import (
    build_feature_context,
    build_model_status,
    load_deployed_model,
    persist_scoring_snapshot,
    score_beneficiary,
)


REGION_COLORS = ["#c2410c", "#0f766e", "#0f172a", "#1d4ed8", "#b45309"]
RETENTION_BREAKDOWN_DIMENSIONS = ("region", "gender", "household_type", "phase")


def _load_beneficiaries(db: Session) -> list[Beneficiary]:
    statement = (
        select(Beneficiary)
        .options(
            selectinload(Beneficiary.program).selectinload(Program.operational_setting),
            selectinload(Beneficiary.program).selectinload(Program.data_policy),
            selectinload(Beneficiary.monitoring_events),
            selectinload(Beneficiary.interventions),
        )
        .order_by(Beneficiary.created_at.desc())
    )
    return list(db.scalars(statement).unique().all())


def _serialize_intervention(intervention: Intervention) -> schemas.InterventionRecord:
    return schemas.InterventionRecord(
        id=intervention.id,
        beneficiary_id=intervention.beneficiary_id,
        beneficiary_name=intervention.beneficiary.full_name,
        action_type=intervention.action_type,
        support_channel=intervention.support_channel,
        protocol_step=intervention.protocol_step,  # type: ignore[arg-type]
        status=intervention.status,
        verification_status=intervention.verification_status,
        assigned_to=intervention.assigned_to,
        assigned_site=intervention.assigned_site,
        due_at=utc_isoformat(intervention.due_at),
        completed_at=utc_isoformat(intervention.completed_at),
        verified_at=utc_isoformat(intervention.verified_at),
        verification_note=intervention.verification_note,
        dismissal_reason=intervention.dismissal_reason,
        attempt_count=intervention.attempt_count,
        source=intervention.source,
        risk_level=intervention.risk_level,
        priority_rank=intervention.priority_rank,
        note=intervention.note,
        successful=intervention.successful,
        soft_signals=_serialize_soft_signals(intervention.beneficiary),
        logged_at=utc_isoformat(intervention.logged_at) or intervention.logged_at.isoformat() + "Z",
    )


def ensure_program_operational_setting(db: Session, program: Program) -> ProgramOperationalSetting:
    """Return the program's operational settings, creating sensible defaults.

    RetainAI cannot assume one universal dropout definition or one universal
    field-capacity profile. This helper guarantees that every program has a
    persisted settings record before queue generation, validation, or
    write-back logic tries to use program-level assumptions.
    """
    setting = db.scalar(
        select(ProgramOperationalSetting).where(ProgramOperationalSetting.program_id == program.id)
    )
    if setting is not None:
        return setting

    label_definition_preset, dropout_inactivity_days, prediction_window_days = default_operational_profile(
        program.program_type
    )
    setting = ProgramOperationalSetting(
        program_id=program.id,
        label_definition_preset=label_definition_preset,
        dropout_inactivity_days=dropout_inactivity_days,
        prediction_window_days=prediction_window_days,
        label_noise_strategy="operational_soft_labels",
        soft_label_weight=0.35,
        silent_transfer_detection_enabled=True,
        tracing_sms_delay_days=3,
        tracing_call_delay_days=7,
        tracing_visit_delay_days=14,
    )
    db.add(setting)
    db.commit()
    db.refresh(setting)
    return setting


def _serialize_program_operational_setting(setting: ProgramOperationalSetting) -> schemas.ProgramOperationalSettingRead:
    return schemas.ProgramOperationalSettingRead(
        id=setting.id,
        program_id=setting.program_id,
        weekly_followup_capacity=setting.weekly_followup_capacity,
        worker_count=setting.worker_count,
        medium_risk_multiplier=setting.medium_risk_multiplier,
        high_risk_share_floor=setting.high_risk_share_floor,
        review_window_days=setting.review_window_days,
        label_definition_preset=setting.label_definition_preset,  # type: ignore[arg-type]
        dropout_inactivity_days=setting.dropout_inactivity_days,
        prediction_window_days=setting.prediction_window_days,
        label_noise_strategy=setting.label_noise_strategy,  # type: ignore[arg-type]
        soft_label_weight=setting.soft_label_weight,
        silent_transfer_detection_enabled=setting.silent_transfer_detection_enabled,
        low_risk_channel=setting.low_risk_channel,  # type: ignore[arg-type]
        medium_risk_channel=setting.medium_risk_channel,  # type: ignore[arg-type]
        high_risk_channel=setting.high_risk_channel,  # type: ignore[arg-type]
        tracing_sms_delay_days=setting.tracing_sms_delay_days,
        tracing_call_delay_days=setting.tracing_call_delay_days,
        tracing_visit_delay_days=setting.tracing_visit_delay_days,
        escalation_window_days=setting.escalation_window_days,
        escalation_max_attempts=setting.escalation_max_attempts,
        fairness_reweighting_enabled=setting.fairness_reweighting_enabled,
        fairness_target_dimensions=list(setting.fairness_target_dimensions or []),
        fairness_max_gap=setting.fairness_max_gap,
        fairness_min_group_size=setting.fairness_min_group_size,
        updated_at=setting.updated_at,
    )


def list_program_operational_settings(db: Session) -> list[schemas.ProgramOperationalSettingRead]:
    programs = db.scalars(select(Program).order_by(Program.name.asc())).all()
    return [_serialize_program_operational_setting(ensure_program_operational_setting(db, program)) for program in programs]


def update_program_operational_setting(
    db: Session,
    program_id: str,
    payload: schemas.ProgramOperationalSettingUpdate,
) -> schemas.ProgramOperationalSettingRead:
    program = db.get(Program, program_id)
    if program is None:
        raise ValueError("Program not found")
    setting = ensure_program_operational_setting(db, program)
    setting.weekly_followup_capacity = payload.weekly_followup_capacity
    setting.worker_count = payload.worker_count
    setting.medium_risk_multiplier = payload.medium_risk_multiplier
    setting.high_risk_share_floor = payload.high_risk_share_floor
    setting.review_window_days = payload.review_window_days
    setting.label_definition_preset = payload.label_definition_preset
    setting.dropout_inactivity_days = payload.dropout_inactivity_days
    setting.prediction_window_days = payload.prediction_window_days
    setting.label_noise_strategy = payload.label_noise_strategy
    setting.soft_label_weight = payload.soft_label_weight
    setting.silent_transfer_detection_enabled = payload.silent_transfer_detection_enabled
    setting.low_risk_channel = payload.low_risk_channel
    setting.medium_risk_channel = payload.medium_risk_channel
    setting.high_risk_channel = payload.high_risk_channel
    setting.tracing_sms_delay_days = payload.tracing_sms_delay_days
    setting.tracing_call_delay_days = payload.tracing_call_delay_days
    setting.tracing_visit_delay_days = payload.tracing_visit_delay_days
    setting.escalation_window_days = payload.escalation_window_days
    setting.escalation_max_attempts = payload.escalation_max_attempts
    setting.fairness_reweighting_enabled = payload.fairness_reweighting_enabled
    setting.fairness_target_dimensions = list(payload.fairness_target_dimensions)
    setting.fairness_max_gap = payload.fairness_max_gap
    setting.fairness_min_group_size = payload.fairness_min_group_size
    db.add(setting)
    db.commit()
    db.refresh(setting)
    return _serialize_program_operational_setting(setting)


def list_interventions(db: Session, limit: int = 10) -> list[schemas.InterventionRecord]:
    statement = (
        select(Intervention)
        .options(selectinload(Intervention.beneficiary))
        .order_by(Intervention.logged_at.desc())
        .limit(limit)
    )
    return [_serialize_intervention(item) for item in db.scalars(statement).all()]


def build_intervention_effectiveness_summary(db: Session) -> schemas.InterventionEffectivenessSummary:
    interventions = list(
        db.scalars(
            select(Intervention)
            .options(selectinload(Intervention.beneficiary).selectinload(Beneficiary.program))
            .order_by(Intervention.logged_at.desc())
        ).all()
    )
    labeled = [item for item in interventions if item.successful is not None]
    grouped: dict[tuple[str, str], list[Intervention]] = defaultdict(list)
    for intervention in labeled:
        context_label = f"{intervention.beneficiary.program.program_type} | {intervention.beneficiary.region}"
        grouped[(intervention.action_type, context_label)].append(intervention)

    rows: list[schemas.InterventionEffectivenessRow] = []
    for (action_type, context_label), items in grouped.items():
        successful = sum(1 for item in items if item.successful)
        attempts = len(items)
        success_rate = round(successful / max(1, attempts), 4)
        avg_risk_score = 0.0
        if items:
            avg_risk_score = round(
                mean(
                    min(
                        100.0,
                        max(
                            0.0,
                            float(build_feature_context(item.beneficiary).last_contact_days / 2),
                        ),
                    )
                    for item in items
                ),
                2,
            )
        strength = "high" if attempts >= 5 and success_rate >= 0.65 else "medium" if attempts >= 3 and success_rate >= 0.45 else "low"
        rows.append(
            schemas.InterventionEffectivenessRow(
                action_type=action_type,
                context_label=context_label,
                attempts=attempts,
                successful_interventions=successful,
                success_rate=success_rate,
                avg_risk_score=avg_risk_score,
                recommendation_strength=strength,  # type: ignore[arg-type]
            )
        )

    rows.sort(key=lambda item: (-item.success_rate, -item.attempts, item.action_type))
    top_recommendations = [
        f"{row.action_type} is performing at {round(row.success_rate * 100)}% success in {row.context_label}."
        for row in rows[:3]
    ] or ["Not enough labeled intervention outcomes yet to recommend an action pattern."]
    narrative = (
        f"{len(labeled)} of {len(interventions)} logged interventions have a recorded outcome. "
        f"The strongest observed pattern is {rows[0].action_type.lower()} in {rows[0].context_label}."
        if rows
        else "Intervention outcomes are not labeled consistently enough yet to estimate effectiveness."
    )
    return schemas.InterventionEffectivenessSummary(
        narrative=narrative,
        total_logged_interventions=len(interventions),
        outcome_labeled_interventions=len(labeled),
        rows=rows[:12],
        top_recommendations=top_recommendations,
    )


def _recommended_intervention_for(
    effectiveness: schemas.InterventionEffectivenessSummary,
    beneficiary: Beneficiary,
    default_action: str,
) -> str:
    program_type = beneficiary.program.program_type
    region = beneficiary.region
    candidates = [
        row
        for row in effectiveness.rows
        if program_type in row.context_label and region in row.context_label and row.recommendation_strength != "low"
    ]
    if not candidates:
        return default_action
    best = max(candidates, key=lambda row: (row.success_rate, row.attempts))
    return f"{best.action_type} (learned best action from {best.attempts} similar interventions)"


def _serialize_soft_signals(beneficiary: Beneficiary) -> schemas.SoftSignalSnapshot | None:
    payload = {
        "household_stability_signal": beneficiary.household_stability_signal,
        "economic_stress_signal": beneficiary.economic_stress_signal,
        "family_support_signal": beneficiary.family_support_signal,
        "health_change_signal": beneficiary.health_change_signal,
        "motivation_signal": beneficiary.motivation_signal,
    }
    if all(value is None for value in payload.values()):
        return None
    return schemas.SoftSignalSnapshot(**payload)


def _select_workflow_intervention(interventions: list[Intervention]) -> Intervention | None:
    ordered = sorted(interventions, key=lambda item: item.logged_at, reverse=True)
    open_intervention = next(
        (
            intervention
            for intervention in ordered
            if intervention.status not in {"closed", "dismissed"}
        ),
        None,
    )
    return open_intervention or (ordered[0] if ordered else None)


def _workflow_state(
    intervention: Intervention | None,
    *,
    risk_level: str,
    beneficiary: Beneficiary,
) -> schemas.FollowUpWorkflowState | None:
    if intervention is None:
        return None
    profile = build_operational_settings_profile(
        beneficiary.program.program_type,
        beneficiary.program.operational_setting,
    )
    projection = project_tracing_protocol(
        risk_level=risk_level,
        profile=profile,
        workflow=intervention,
    )
    return schemas.FollowUpWorkflowState(
        intervention_id=intervention.id,
        status=intervention.status,  # type: ignore[arg-type]
        verification_status=intervention.verification_status,  # type: ignore[arg-type]
        assigned_to=intervention.assigned_to,
        assigned_site=intervention.assigned_site,
        due_at=intervention.due_at,
        completed_at=intervention.completed_at,
        verified_at=intervention.verified_at,
        note=intervention.note,
        verification_note=intervention.verification_note,
        dismissal_reason=intervention.dismissal_reason,
        support_channel=intervention.support_channel,
        protocol_step=(intervention.protocol_step or projection.current_step),  # type: ignore[arg-type]
        attempt_count=intervention.attempt_count,
        successful=intervention.successful,
        tracing_protocol=schemas.TracingProtocolState(
            current_step=projection.current_step,  # type: ignore[arg-type]
            current_channel=projection.current_channel,  # type: ignore[arg-type]
            current_due_at=projection.current_due_at,
            next_step=projection.next_step,  # type: ignore[arg-type]
            next_due_at=projection.next_due_at,
            sms_delay_days=profile.tracing_sms_delay_days,
            call_delay_days=profile.tracing_call_delay_days,
            visit_delay_days=profile.tracing_visit_delay_days,
        ),
    )


def _workflow_status_label(intervention: Intervention | None) -> str:
    if intervention is None:
        return "Not contacted"
    if intervention.status == "dismissed":
        return f"Dismissed{f': {intervention.dismissal_reason}' if intervention.dismissal_reason else ''}"
    if intervention.status == "closed":
        return (
            f"Closed: {intervention.verification_status.replace('_', ' ')}"
            if intervention.verification_status
            else "Closed"
        )
    if intervention.status == "verified":
        return (
            f"Verified: {intervention.verification_status.replace('_', ' ')}"
            if intervention.verification_status
            else "Verified"
        )
    if intervention.status == "reached":
        return "Reached - verify status"
    if intervention.status == "attempted":
        return f"Attempted ({intervention.attempt_count} tries)"
    if intervention.status == "escalated":
        return "Escalated for supervisor review"
    if intervention.status == "queued":
        return f"Queued for {intervention.assigned_to or 'follow-up queue'}"
    return intervention.status.replace("_", " ").title()


def _apply_program_capacity_thresholds(
    db: Session,
    cases: list[schemas.RiskCase],
    beneficiaries_by_id: dict[str, Beneficiary],
) -> list[schemas.RiskCase]:
    grouped: dict[str, list[schemas.RiskCase]] = defaultdict(list)
    for case in cases:
        grouped[case.program].append(case)

    adjusted: list[schemas.RiskCase] = []
    for group_cases in grouped.values():
        if not group_cases:
            continue
        first_beneficiary = beneficiaries_by_id.get(group_cases[0].id)
        if first_beneficiary is None:
            adjusted.extend(group_cases)
            continue
        setting = ensure_program_operational_setting(db, first_beneficiary.program)
        high_count = max(
            1,
            min(
                len(group_cases),
                max(
                    int(ceil(len(group_cases) * setting.high_risk_share_floor)),
                    setting.weekly_followup_capacity,
                ),
            ),
        )
        medium_count = min(
            len(group_cases),
            max(high_count, int(ceil(setting.weekly_followup_capacity * setting.medium_risk_multiplier))),
        )
        sorted_cases = sorted(group_cases, key=lambda item: (-item.risk_score, item.name))
        worker_count = max(1, setting.worker_count)
        per_worker_capacity = max(1, int(ceil(setting.weekly_followup_capacity / worker_count)))
        worker_site_ranks: dict[tuple[str, str], int] = defaultdict(int)
        for index, case in enumerate(sorted_cases):
            level = "Low"
            if index < high_count:
                level = "High"
            elif index < medium_count:
                level = "Medium"
            case.risk_level = level  # type: ignore[assignment]
            beneficiary = beneficiaries_by_id.get(case.id)
            assigned_site = case.assigned_site or (beneficiary.assigned_site if beneficiary is not None else None) or case.region
            assigned_worker = case.assigned_worker or (
                beneficiary.assigned_case_worker if beneficiary is not None else None
            )
            if not assigned_worker:
                assigned_worker = f"{case.program} queue {(index % worker_count) + 1}"
            case.assigned_worker = assigned_worker
            case.assigned_site = assigned_site
            group_key = (assigned_worker, assigned_site)
            worker_site_ranks[group_key] += 1
            case.queue_rank = worker_site_ranks[group_key]
            if case.queue_rank <= per_worker_capacity:
                case.queue_bucket = "Due now"
            elif case.queue_rank <= per_worker_capacity * 2:
                case.queue_bucket = "This week"
            else:
                case.queue_bucket = "Monitor"
            adjusted.append(case)
    bucket_priority = {"Due now": 0, "This week": 1, "Monitor": 2}
    return sorted(adjusted, key=lambda item: (bucket_priority[item.queue_bucket], item.queue_rank, -item.risk_score, item.name))


def build_risk_cases(db: Session) -> list[schemas.RiskCase]:
    """Build the ranked operational queue consumed by the dashboard.

    This function is one of the most product-critical points in the backend. It
    combines beneficiary history, model output, heuristic fallbacks, governance
    state, intervention context, and capacity logic into a single user-facing
    case record per beneficiary.
    """
    beneficiaries = _load_beneficiaries(db)
    active_like_statuses = {"active", "at_risk", "enrolled"}
    cases: list[schemas.RiskCase] = []
    beneficiaries_by_id = {item.id: item for item in beneficiaries}
    loaded_model = load_deployed_model(db)
    intervention_effectiveness = build_intervention_effectiveness_summary(db)

    for beneficiary in beneficiaries:
        if beneficiary.status not in active_like_statuses:
            continue

        model_prediction, heuristic = score_beneficiary(beneficiary, None if beneficiary.opted_out else loaded_model)
        profile = build_operational_settings_profile(
            beneficiary.program.program_type,
            beneficiary.program.operational_setting,
        )
        confidence = model_prediction.confidence
        if beneficiary.opted_out:
            confidence = "Opted out of modeling"

        feature_context = build_feature_context(beneficiary)
        persist_scoring_snapshot(
            db,
            beneficiary,
            model_version=None if beneficiary.opted_out or loaded_model is None else loaded_model.version,
            feature_values=feature_context.features,
            uncertainty_score=model_prediction.uncertainty_score,
        )
        workflow_intervention = _select_workflow_intervention(list(beneficiary.interventions))
        tracing_projection = project_tracing_protocol(
            risk_level=model_prediction.risk_level,
            profile=profile,
            workflow=workflow_intervention,
        )

        cases.append(
            schemas.RiskCase(
                id=beneficiary.id,
                name=beneficiary.full_name,
                program=beneficiary.program.name,
                program_type=beneficiary.program.program_type,
                region=beneficiary.region,
                cohort=beneficiary.cohort,
                phase=beneficiary.phase,
                risk_level=model_prediction.risk_level,  # type: ignore[arg-type]
                risk_score=model_prediction.risk_score,
                explanation=model_prediction.explanation,
                recommended_action=_recommended_intervention_for(
                    intervention_effectiveness,
                    beneficiary,
                    tracing_recommended_action(beneficiary.program.program_type, tracing_projection),
                ),
                flags=model_prediction.flags,
                last_contact_days=heuristic.last_contact_days,
                attendance_rate_30d=heuristic.attendance_rate_30d,
                intervention_status=_workflow_status_label(workflow_intervention),
                confidence=confidence,  # type: ignore[arg-type]
                opted_out=beneficiary.opted_out,
                assigned_worker=(
                    workflow_intervention.assigned_to
                    if workflow_intervention is not None and workflow_intervention.assigned_to
                    else beneficiary.assigned_case_worker
                ),
                assigned_site=(
                    workflow_intervention.assigned_site
                    if workflow_intervention is not None and workflow_intervention.assigned_site
                    else beneficiary.assigned_site
                ),
                queue_rank=0,
                queue_bucket="Monitor",
                workflow=_workflow_state(
                    workflow_intervention,
                    risk_level=model_prediction.risk_level,
                    beneficiary=beneficiary,
                ),
                tracing_protocol=schemas.TracingProtocolState(
                    current_step=tracing_projection.current_step,  # type: ignore[arg-type]
                    current_channel=tracing_projection.current_channel,  # type: ignore[arg-type]
                    current_due_at=tracing_projection.current_due_at,
                    next_step=tracing_projection.next_step,  # type: ignore[arg-type]
                    next_due_at=tracing_projection.next_due_at,
                    sms_delay_days=profile.tracing_sms_delay_days,
                    call_delay_days=profile.tracing_call_delay_days,
                    visit_delay_days=profile.tracing_visit_delay_days,
                ),
                soft_signals=_serialize_soft_signals(beneficiary),
            )
        )

    if beneficiaries:
        db.commit()

    return _apply_program_capacity_thresholds(db, cases, beneficiaries_by_id)


def filter_risk_cases(
    cases: list[schemas.RiskCase],
    *,
    program: str | None = None,
    risk_level: str | None = None,
    region: str | None = None,
    cohort: str | None = None,
    phase: str | None = None,
    search: str | None = None,
) -> list[schemas.RiskCase]:
    """Apply dashboard and export filters to already-built queue cases.

    Filtering is kept separate from queue construction so the same underlying
    ranked case list can drive:

    - the interactive dashboard
    - CSV or messaging exports
    - connector write-back payload generation
    """
    filtered = cases

    if program:
        program_needle = program.strip().lower()
        filtered = [case for case in filtered if case.program.lower() == program_needle]

    if risk_level:
        filtered = [case for case in filtered if case.risk_level == risk_level]

    if region:
        region_needle = region.strip().lower()
        filtered = [case for case in filtered if case.region.lower() == region_needle]

    if cohort:
        cohort_needle = cohort.strip().lower()
        filtered = [case for case in filtered if (case.cohort or "").lower() == cohort_needle]

    if phase:
        phase_needle = phase.strip().lower()
        filtered = [case for case in filtered if (case.phase or "").lower() == phase_needle]

    if search:
        needle = search.strip().lower()
        filtered = [
            case
            for case in filtered
            if needle in case.name.lower()
            or needle in case.program.lower()
            or needle in case.region.lower()
            or needle in (case.cohort or "").lower()
            or needle in (case.phase or "").lower()
        ]

    return filtered


def build_follow_up_export(
    db: Session,
    *,
    cases: list[schemas.RiskCase],
    mode: str,
) -> str:
    """Render a queue slice as a channel-specific CSV export.

    The output is intentionally lightweight and operational rather than generic.
    Each export mode packages the same risk cases in a shape that is easier to
    use for a specific follow-up channel such as WhatsApp, SMS, or field visits.
    """
    beneficiaries = {
        beneficiary.id: beneficiary
        for beneficiary in _load_beneficiaries(db)
    }

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "beneficiary_name",
            "assigned_to",
            "assigned_site",
            "queue_bucket",
            "queue_rank",
            "channel",
            "phone",
            "program",
            "region",
            "cohort",
            "phase",
            "risk_level",
            "risk_score",
            "recommended_action",
            "message",
        ],
    )
    writer.writeheader()

    for case in cases:
        beneficiary = beneficiaries.get(case.id)
        phone = beneficiary.preferred_contact_phone if beneficiary is not None else ""
        channel = (
            beneficiary.preferred_contact_channel
            if beneficiary is not None and beneficiary.preferred_contact_channel
            else ("field_visit" if mode == "field_visit" else mode)
        )
        if mode == "whatsapp":
            message = (
                f"Hello {case.name.split(' ')[0]}, we would like to check in and support your participation in {case.program}. "
                f"Our team noted that a follow-up may help before the next scheduled activity."
            )
        elif mode == "sms":
            message = (
                f"{case.program}: our team would like to support your continued participation. "
                f"Please expect a follow-up regarding your next step."
            )
        else:
            message = (
                f"Visit {case.name} in {case.region}. Review barrier signals: {', '.join(case.flags[:2]) or 'general check-in'}."
            )

        writer.writerow(
            {
                "beneficiary_name": case.name,
                "assigned_to": case.assigned_worker or "",
                "assigned_site": case.assigned_site or "",
                "queue_bucket": case.queue_bucket,
                "queue_rank": case.queue_rank,
                "channel": channel,
                "phone": phone or "",
                "program": case.program,
                "region": case.region,
                "cohort": case.cohort or "",
                "phase": case.phase or "",
                "risk_level": case.risk_level,
                "risk_score": case.risk_score,
                "recommended_action": case.recommended_action,
                "message": message,
            }
        )

    return output.getvalue()


def build_dashboard_summary(db: Session, cases: list[schemas.RiskCase] | None = None) -> schemas.DashboardSummary:
    """Assemble the high-level summary cards shown on the dashboard shell.

    The summary is a compact portfolio view over the richer queue and analytics
    data. It is optimized for quick supervisory scanning rather than detailed
    investigation.
    """
    risk_cases = cases if cases is not None else build_risk_cases(db)
    beneficiaries = _load_beneficiaries(db)
    model_status = build_model_status(db)
    active_beneficiaries = sum(1 for beneficiary in beneficiaries if beneficiary.status in {"active", "at_risk", "enrolled"})
    high_risk_cases = [case for case in risk_cases if case.risk_level == "High"]
    medium_risk_cases = [case for case in risk_cases if case.risk_level == "Medium"]

    predicted_30_day_dropout = round(
        sum(case.risk_score / 100 for case in risk_cases if case.risk_level != "Low")
    )

    interventions = db.scalars(select(Intervention)).all()
    completed_interventions = [intervention for intervention in interventions if intervention.successful is not None]
    if completed_interventions:
        intervention_success_rate = round(
            (sum(1 for intervention in completed_interventions if intervention.successful) / len(completed_interventions)) * 100
        )
    else:
        intervention_success_rate = 0

    driver_counter: Counter[str] = Counter()
    for case in high_risk_cases + medium_risk_cases:
        driver_counter.update(case.flags[:3])

    top_risk_drivers = [
        schemas.TopRiskDriver(
            name=driver_name,
            impacted_beneficiaries=count,
            insight=f"{count} active beneficiaries are currently being elevated by this signal.",
        )
        for driver_name, count in driver_counter.most_common(4)
    ]

    if not top_risk_drivers:
        top_risk_drivers = [
            schemas.TopRiskDriver(
                name="No current high-risk drivers",
                impacted_beneficiaries=0,
                insight="No active beneficiaries have reached elevated risk thresholds yet.",
            )
        ]

    retention_curves = build_retention_curves(db)
    region_alerts = _build_region_alerts(retention_curves, risk_cases)

    return schemas.DashboardSummary(
        active_beneficiaries=active_beneficiaries,
        high_risk_cases=len(high_risk_cases),
        predicted_30_day_dropout=predicted_30_day_dropout,
        intervention_success_rate=intervention_success_rate,
        weekly_followups_due=sum(1 for case in risk_cases if case.queue_bucket != "Monitor"),
        model_mode=model_status.model_mode,
        last_retrained=(model_status.trained_at or date.today().isoformat())[:10],
        quality_note=model_status.notes or "Risk scores are generated from live engagement, vulnerability, and event-history data stored in the platform database.",
        top_risk_drivers=top_risk_drivers,
        region_alerts=region_alerts,
        model_status=model_status,
    )


def build_retention_curves(db: Session) -> schemas.RetentionCurves:
    """Build plain-language cohort retention curves for the analytics view.

    The implementation is intentionally simple and explainable. It aims to give
    program users a usable sense of retention shape by region and month without
    requiring statistical literacy or survival-analysis jargon.
    """
    beneficiaries = _load_beneficiaries(db)
    if not beneficiaries:
        return schemas.RetentionCurves(
            narrative="No beneficiary data has been loaded yet.",
            series=[],
            data=[],
        )

    region_counter = Counter(beneficiary.region for beneficiary in beneficiaries)
    top_regions = [region for region, _ in region_counter.most_common(3)]
    region_beneficiaries = {region: [beneficiary for beneficiary in beneficiaries if beneficiary.region == region] for region in top_regions}

    series: list[schemas.RetentionSeries] = []
    for index, region in enumerate(top_regions):
        key = region.lower().replace(" ", "_").replace("-", "_")
        series.append(
            schemas.RetentionSeries(
                key=key,
                label=region,
                color=REGION_COLORS[index % len(REGION_COLORS)],
            )
        )

    max_month = 6
    data: list[dict[str, str | float]] = []

    for month_index in range(max_month + 1):
        label = "Enrollment" if month_index == 0 else f"Month {month_index}"
        row: dict[str, str | float] = {"period": label}

        for item in series:
            beneficiaries_for_region = region_beneficiaries[item.label]
            retained = 0
            for beneficiary in beneficiaries_for_region:
                checkpoint = beneficiary.enrollment_date + timedelta(days=30 * month_index)
                if beneficiary.dropout_date is None or beneficiary.dropout_date > checkpoint:
                    retained += 1
            retention_pct = round((retained / len(beneficiaries_for_region)) * 100, 1) if beneficiaries_for_region else 0.0
            row[item.key] = retention_pct

        data.append(row)

    region_deltas = {
        item.label: float(data[-1][item.key]) - float(data[max(0, len(data) - 4)][item.key])
        for item in series
    }
    worst_region = min(region_deltas, key=region_deltas.get)
    best_region = max(region_deltas, key=region_deltas.get)
    narrative = (
        f"{worst_region} retention moved {region_deltas[worst_region]:.1f} points over the last three monthly checkpoints, "
        f"while {best_region} moved {region_deltas[best_region]:.1f} points."
    )

    return schemas.RetentionCurves(
        narrative=narrative,
        series=series,
        data=data,
    )


def build_retention_analytics(db: Session) -> schemas.RetentionAnalytics:
    """Build portfolio-level retention breakdowns and trend signals.

    This function turns beneficiary outcome history into the aggregate views used
    for internal program management and donor-facing reporting.
    """
    beneficiaries = _load_beneficiaries(db)
    curves = build_retention_curves(db)
    breakdowns: list[schemas.RetentionBreakdownRow] = []

    for dimension in RETENTION_BREAKDOWN_DIMENSIONS:
        grouped: dict[str, list[Beneficiary]] = defaultdict(list)
        for beneficiary in beneficiaries:
            if dimension == "phase":
                key = beneficiary.phase or "Unknown"
            else:
                key = str(getattr(beneficiary, dimension) or "Unknown")
            grouped[key].append(beneficiary)

        for group_name, items in grouped.items():
            active = sum(1 for item in items if item.status in {"active", "at_risk", "enrolled", "completed"})
            dropped = sum(1 for item in items if item.status == "dropped")
            recent_dropped = sum(
                1
                for item in items
                if item.dropout_date is not None and (date.today() - item.dropout_date).days <= 90
            )
            total = len(items)
            breakdowns.append(
                schemas.RetentionBreakdownRow(
                    dimension=dimension,
                    group_name=group_name,
                    active_beneficiaries=active,
                    dropped_beneficiaries=dropped,
                    retention_rate=round(active / max(1, total), 4),
                    recent_dropout_rate=round(recent_dropped / max(1, total), 4),
                )
            )

    breakdowns.sort(key=lambda item: (item.dimension, item.retention_rate, item.group_name))

    periods = [date.today().replace(day=1) - timedelta(days=30 * index) for index in range(5)]
    trend_rows: list[schemas.RetentionTrendRow] = []
    for period in sorted(periods):
        label = period.strftime("%Y-%m")
        for region in sorted({item.region for item in beneficiaries}):
            regional = [item for item in beneficiaries if item.region == region]
            active = sum(
                1
                for item in regional
                if item.enrollment_date <= period and (item.dropout_date is None or item.dropout_date > period)
            )
            dropped = sum(
                1
                for item in regional
                if item.dropout_date is not None and item.dropout_date <= period
            )
            total = max(1, len([item for item in regional if item.enrollment_date <= period]))
            representative = regional[0].program.name if regional else "No program"
            trend_rows.append(
                schemas.RetentionTrendRow(
                    period=label,
                    program_name=representative,
                    region=region,
                    retention_rate=round(active / total, 4),
                    dropout_rate=round(dropped / total, 4),
                    active_beneficiaries=active,
                )
            )

    highlights: list[str] = []
    if trend_rows:
        latest_by_region = defaultdict(list)
        for row in trend_rows:
            latest_by_region[row.region].append(row)
        for region, items in latest_by_region.items():
            ordered = sorted(items, key=lambda item: item.period)
            if len(ordered) >= 2:
                delta = round((ordered[-1].retention_rate - ordered[-2].retention_rate) * 100, 1)
                direction = "declined" if delta < 0 else "improved"
                highlights.append(f"Retention in {region} has {direction} {abs(delta)} points since the last monthly checkpoint.")

    return schemas.RetentionAnalytics(
        narrative=curves.narrative,
        breakdowns=breakdowns[:24],
        trend_rows=trend_rows[-18:],
        trend_highlights=highlights[:6],
        retention_curves=curves,
    )


def build_donor_report_summary(db: Session) -> schemas.DonorReportSummary:
    """Package the main donor-facing narrative and headline metrics.

    The backend keeps donor-report assembly centralized so that the PDF, Excel,
    and API summary views all share the same numbers and narrative framing.
    """
    retention_analytics = build_retention_analytics(db)
    intervention_effectiveness = build_intervention_effectiveness_summary(db)
    dashboard = build_dashboard_summary(db)
    narrative = (
        f"RetainAI is currently monitoring {dashboard.active_beneficiaries} active beneficiaries across "
        f"{len({row.program_name for row in retention_analytics.trend_rows}) or 1} active portfolio slices. "
        f"{dashboard.high_risk_cases} beneficiaries currently fall inside each program's high-priority follow-up capacity."
    )
    return schemas.DonorReportSummary(
        generated_at=utc_isoformat(),
        narrative=narrative,
        headline_metrics={
            "active_beneficiaries": dashboard.active_beneficiaries,
            "high_risk_cases": dashboard.high_risk_cases,
            "predicted_30_day_dropout": dashboard.predicted_30_day_dropout,
            "intervention_success_rate": dashboard.intervention_success_rate,
            "weekly_followups_due": dashboard.weekly_followups_due,
            "model_algorithm": dashboard.model_status.algorithm,
        },
        retention_analytics=retention_analytics,
        intervention_effectiveness=intervention_effectiveness,
    )


def build_donor_excel_report(db: Session) -> bytes:
    summary = build_donor_report_summary(db)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview["A1"] = "RetainAI Donor Summary"
    overview["A1"].font = Font(bold=True, size=14)
    overview["A3"] = summary.narrative
    row_index = 5
    for key, value in summary.headline_metrics.items():
        overview.cell(row=row_index, column=1, value=key)
        overview.cell(row=row_index, column=2, value=value)
        row_index += 1

    breakdown_sheet = workbook.create_sheet("Retention Breakdowns")
    breakdown_sheet.append(["dimension", "group_name", "active_beneficiaries", "dropped_beneficiaries", "retention_rate", "recent_dropout_rate"])
    for row in summary.retention_analytics.breakdowns:
        breakdown_sheet.append([row.dimension, row.group_name, row.active_beneficiaries, row.dropped_beneficiaries, row.retention_rate, row.recent_dropout_rate])

    intervention_sheet = workbook.create_sheet("Interventions")
    intervention_sheet.append(["action_type", "context_label", "attempts", "successful_interventions", "success_rate", "avg_risk_score"])
    for row in summary.intervention_effectiveness.rows:
        intervention_sheet.append([row.action_type, row.context_label, row.attempts, row.successful_interventions, row.success_rate, row.avg_risk_score])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_donor_pdf_report(db: Session) -> bytes:
    summary = build_donor_report_summary(db)
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 50
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(40, y, "RetainAI Donor Summary")
    y -= 26
    pdf.setFont("Helvetica", 10)
    for line in [summary.generated_at, summary.narrative, *summary.retention_analytics.trend_highlights[:4], *summary.intervention_effectiveness.top_recommendations[:3]]:
        if y < 60:
            pdf.showPage()
            y = height - 50
            pdf.setFont("Helvetica", 10)
        pdf.drawString(40, y, str(line)[:110])
        y -= 16
    pdf.showPage()
    pdf.save()
    return output.getvalue()


def _build_region_alerts(
    retention_curves: schemas.RetentionCurves,
    risk_cases: list[schemas.RiskCase],
) -> list[schemas.RegionAlert]:
    alerts: list[schemas.RegionAlert] = []
    regional_drivers: dict[str, Counter[str]] = defaultdict(Counter)

    for case in risk_cases:
        regional_drivers[case.region].update(case.flags[:2])

    for series in retention_curves.series:
        if len(retention_curves.data) < 4:
            break

        latest = float(retention_curves.data[-1][series.key])
        prior = float(retention_curves.data[-4][series.key])
        delta = round(latest - prior, 1)
        top_driver = regional_drivers[series.label].most_common(1)
        driver_text = top_driver[0][0].lower() if top_driver else "recent engagement changes"
        alerts.append(
            schemas.RegionAlert(
                region=series.label,
                retention_delta=delta,
                note=f"{series.label} is being shaped primarily by {driver_text}.",
            )
        )

    return alerts
