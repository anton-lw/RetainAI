"""File ingestion, schema inference, and ETL validation logic.

This module handles the practical messiness of NGO data ingestion: CSV/XLSX
parsing, field mapping, type coercion, anomaly detection, and idempotent import
behavior. It exists so data-quality failures surface explicitly instead of
silently corrupting downstream model behavior.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import schemas
from app.models import Beneficiary, DataQualityIssue, ImportBatch, MonitoringEvent, Program


DatasetType = Literal["beneficiaries", "events"]
SourceFormat = Literal["csv", "xlsx"]


BENEFICIARY_ALIASES: dict[str, list[str]] = {
    "external_id": ["external_id", "beneficiary_id", "participant_id", "id"],
    "full_name": ["full_name", "beneficiary_name", "participant_name", "name"],
    "gender": ["gender", "sex"],
    "region": ["region", "district", "location", "county"],
    "cohort": ["cohort", "group", "wave"],
    "phase": ["phase", "program_phase", "stage"],
    "household_type": ["household_type", "household_category"],
    "delivery_modality": ["delivery_modality", "modality"],
    "enrollment_date": ["enrollment_date", "start_date", "date_enrolled"],
    "dropout_date": ["dropout_date", "exit_date"],
    "completion_date": ["completion_date", "completed_at"],
    "status": ["status", "program_status"],
    "household_size": ["household_size", "hh_size"],
    "pmt_score": ["pmt_score", "poverty_score", "proxy_means_test"],
    "food_insecurity_index": ["food_insecurity_index", "food_insecurity", "hfias_score", "hdds_score"],
    "distance_to_service_km": ["distance_to_service_km", "distance_km", "travel_distance_km"],
    "preferred_contact_phone": ["preferred_contact_phone", "phone", "phone_number", "mobile_number", "whatsapp_number"],
    "preferred_contact_channel": ["preferred_contact_channel", "contact_channel", "preferred_channel"],
    "assigned_case_worker": ["assigned_case_worker", "case_worker", "chw_name", "field_worker", "assigned_worker"],
    "assigned_site": ["assigned_site", "site", "facility", "school", "service_point"],
    "current_note": ["current_note", "field_note", "notes", "last_note"],
    "household_stability_signal": ["household_stability_signal", "household_stability", "soft_household_stability"],
    "economic_stress_signal": ["economic_stress_signal", "economic_stress", "soft_economic_stress"],
    "family_support_signal": ["family_support_signal", "family_support", "soft_family_support"],
    "health_change_signal": ["health_change_signal", "health_change", "soft_health_change"],
    "motivation_signal": ["motivation_signal", "motivation", "soft_motivation"],
    "opted_out": ["opted_out", "model_opt_out"],
    "modeling_consent_status": ["modeling_consent_status", "consent_status", "model_consent"],
    "consent_method": ["consent_method", "consent_channel"],
    "consent_note": ["consent_note", "consent_comments"],
}

EVENT_ALIASES: dict[str, list[str]] = {
    "external_id": ["external_id", "beneficiary_id", "participant_id", "id"],
    "event_date": ["event_date", "date", "visit_date", "checkin_date"],
    "event_type": ["event_type", "type", "interaction_type", "activity_type"],
    "successful": ["successful", "attended", "present", "was_successful", "outcome"],
    "response_received": ["response_received", "responded", "response"],
    "source": ["source", "channel"],
    "notes": ["notes", "note", "field_note", "comments"],
}

REQUIRED_FIELDS: dict[DatasetType, set[str]] = {
    "beneficiaries": {"external_id", "full_name", "enrollment_date", "region"},
    "events": {"external_id", "event_date", "event_type"},
}

DATE_FIELDS = {"enrollment_date", "dropout_date", "completion_date", "event_date"}
NUMERIC_FIELDS = {
    "household_size",
    "pmt_score",
    "food_insecurity_index",
    "distance_to_service_km",
    "household_stability_signal",
    "economic_stress_signal",
    "family_support_signal",
    "health_change_signal",
    "motivation_signal",
}
BOOLEAN_FIELDS = {"opted_out", "successful", "response_received"}


@dataclass
class ImportAnalysis:
    dataset_type: DatasetType
    source_format: SourceFormat
    records_received: int
    duplicate_rows: int
    inferred_types: dict[str, str]
    suggested_mapping: dict[str, str | None]
    quality_score: int
    warnings: list[str]
    issues: list[schemas.DataQualityIssueRecord]
    sample_rows: list[dict[str, str]]
    unique_rows: list[dict[str, str]]


def _normalize_header(value: str) -> str:
    return "".join(character.lower() if character.isalnum() else "_" for character in value).strip("_")


def detect_mapping(headers: list[str], dataset_type: DatasetType) -> dict[str, str | None]:
    aliases = BENEFICIARY_ALIASES if dataset_type == "beneficiaries" else EVENT_ALIASES
    normalized = {_normalize_header(header): header for header in headers}
    mapping: dict[str, str | None] = {}

    for field_name, options in aliases.items():
        mapping[field_name] = None
        for option in options:
            if option in normalized:
                mapping[field_name] = normalized[option]
                break

    return mapping


def parse_csv_bytes(file_bytes: bytes) -> list[dict[str, str]]:
    decoded = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))
    return [
        {str(key): "" if value is None else str(value) for key, value in row.items()}
        for row in reader
        if any((value or "").strip() for value in row.values())
    ]


def parse_tabular_bytes(file_bytes: bytes, filename: str) -> tuple[list[dict[str, str]], SourceFormat]:
    extension = Path(filename).suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], "xlsx"
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        parsed_rows: list[dict[str, str]] = []
        for values in rows[1:]:
            row = {
                headers[index]: "" if value is None else str(value).strip()
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            if any(value.strip() for value in row.values()):
                parsed_rows.append(row)
        return parsed_rows, "xlsx"
    return parse_csv_bytes(file_bytes), "csv"


def _read_value(row: dict[str, str], column_name: str | None) -> str | None:
    if not column_name:
        return None
    raw = row.get(column_name)
    if raw is None:
        return None
    value = raw.strip()
    return value or None


def _parse_date(value: str | None) -> date | None:
    if value is None:
        return None

    formats = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S")
    for date_format in formats:
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value}")


def _parse_int(value: str | None) -> int | None:
    if value is None:
        return None
    return int(float(value))


def _parse_float(value: str | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _parse_bool(value: str | None) -> bool | None:
    if value is None:
        return None

    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y", "present", "attended", "completed", "successful"}:
        return True
    if normalized in {"0", "false", "no", "n", "missed", "absent", "failed", "unsuccessful"}:
        return False
    return None


def validate_mapping(mapping: dict[str, str | None], dataset_type: DatasetType) -> list[str]:
    return [field_name for field_name in REQUIRED_FIELDS[dataset_type] if not mapping.get(field_name)]


def _infer_scalar_type(value: str) -> str:
    if _parse_bool(value) is not None:
        return "boolean"
    try:
        _parse_date(value)
    except ValueError:
        pass
    else:
        return "date"
    try:
        int(float(value))
    except ValueError:
        pass
    else:
        return "integer"
    try:
        float(value)
    except ValueError:
        return "text"
    return "float"


def infer_column_types(rows: list[dict[str, str]]) -> dict[str, str]:
    if not rows:
        return {}

    inferred: dict[str, str] = {}
    for header in rows[0].keys():
        values = [row.get(header, "").strip() for row in rows[:100] if row.get(header, "").strip()]
        if not values:
            inferred[header] = "empty"
            continue
        observed = {_infer_scalar_type(value) for value in values}
        if observed == {"boolean"}:
            inferred[header] = "boolean"
        elif observed <= {"integer"}:
            inferred[header] = "integer"
        elif observed <= {"integer", "float"}:
            inferred[header] = "float"
        elif observed == {"date"}:
            inferred[header] = "date"
        else:
            inferred[header] = "text"
    return inferred


def _build_issue(
    *,
    severity: Literal["info", "warning", "error"],
    issue_type: str,
    message: str,
    field_name: str | None = None,
    row_number: int | None = None,
    sample_value: str | None = None,
) -> schemas.DataQualityIssueRecord:
    return schemas.DataQualityIssueRecord(
        severity=severity,
        issue_type=issue_type,
        field_name=field_name,
        row_number=row_number,
        message=message,
        sample_value=sample_value,
    )


def _duplicate_key(row: dict[str, str], dataset_type: DatasetType, mapping: dict[str, str | None]) -> tuple[str, ...] | None:
    external_id = _read_value(row, mapping.get("external_id"))
    if not external_id:
        return None
    if dataset_type == "beneficiaries":
        return (external_id,)
    event_date = _read_value(row, mapping.get("event_date"))
    event_type = _read_value(row, mapping.get("event_type"))
    if not event_date or not event_type:
        return None
    return (external_id, event_date, event_type.lower())


def _row_anomalies(
    row: dict[str, str],
    row_number: int,
    dataset_type: DatasetType,
    mapping: dict[str, str | None],
) -> list[schemas.DataQualityIssueRecord]:
    issues: list[schemas.DataQualityIssueRecord] = []

    for field_name in REQUIRED_FIELDS[dataset_type]:
        value = _read_value(row, mapping.get(field_name))
        if not value:
            issues.append(
                _build_issue(
                    severity="error",
                    issue_type="missing_required",
                    field_name=field_name,
                    row_number=row_number,
                    message=f"Missing required value for {field_name}.",
                )
            )

    for field_name in DATE_FIELDS.intersection(mapping.keys()):
        raw = _read_value(row, mapping.get(field_name))
        if raw is None:
            continue
        try:
            _parse_date(raw)
        except ValueError:
            issues.append(
                _build_issue(
                    severity="error",
                    issue_type="invalid_date",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message=f"Unsupported date format for {field_name}.",
                )
            )

    for field_name in BOOLEAN_FIELDS.intersection(mapping.keys()):
        raw = _read_value(row, mapping.get(field_name))
        if raw is not None and _parse_bool(raw) is None:
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="ambiguous_boolean",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message=f"Boolean field {field_name} could not be parsed confidently.",
                )
            )

    for field_name in NUMERIC_FIELDS.intersection(mapping.keys()):
        raw = _read_value(row, mapping.get(field_name))
        if raw is None:
            continue
        try:
            numeric_value = _parse_float(raw)
        except ValueError:
            issues.append(
                _build_issue(
                    severity="error",
                    issue_type="invalid_numeric",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message=f"Numeric field {field_name} could not be parsed.",
                )
            )
            continue
        assert numeric_value is not None

        if field_name == "household_size" and (numeric_value < 0 or numeric_value > 25):
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="outlier_household_size",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message="Household size falls outside the expected 0-25 range.",
                )
            )
        if field_name == "distance_to_service_km" and (numeric_value < 0 or numeric_value > 150):
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="outlier_distance",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message="Distance to service point falls outside the expected 0-150 km range.",
                )
            )
        if field_name == "pmt_score" and (numeric_value < 0 or numeric_value > 100):
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="outlier_pmt",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message="PMT score falls outside the expected 0-100 range.",
                )
            )
        if field_name == "food_insecurity_index" and (numeric_value < 0 or numeric_value > 10):
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="outlier_food_insecurity",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message="Food insecurity score falls outside the expected 0-10 range.",
                )
            )
        if field_name in {
            "household_stability_signal",
            "economic_stress_signal",
            "family_support_signal",
            "health_change_signal",
            "motivation_signal",
        } and (numeric_value < 1 or numeric_value > 5):
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="outlier_soft_signal",
                    field_name=field_name,
                    row_number=row_number,
                    sample_value=raw,
                    message="Soft-indicator values should usually be recorded on a 1-5 scale.",
                )
            )

    return issues


def analyze_rows(
    rows: list[dict[str, str]],
    *,
    dataset_type: DatasetType,
    source_format: SourceFormat,
    mapping: dict[str, str | None],
) -> ImportAnalysis:
    issues: list[schemas.DataQualityIssueRecord] = []
    deduplicated: dict[tuple[str, ...], dict[str, str]] = {}
    duplicate_rows = 0

    for row_number, row in enumerate(rows, start=2):
        issues.extend(_row_anomalies(row, row_number, dataset_type, mapping))
        key = _duplicate_key(row, dataset_type, mapping)
        if key is None:
            continue
        if key in deduplicated:
            duplicate_rows += 1
            issues.append(
                _build_issue(
                    severity="warning",
                    issue_type="duplicate_row",
                    row_number=row_number,
                    message="Duplicate record key detected in uploaded file; latest row was retained.",
                    sample_value=" | ".join(key),
                )
            )
        deduplicated[key] = row

    warnings = [issue.message for issue in issues if issue.severity != "info"][:10]
    error_count = sum(1 for issue in issues if issue.severity == "error")
    warning_count = sum(1 for issue in issues if issue.severity == "warning")
    quality_score = max(0, min(100, 100 - (error_count * 8) - (warning_count * 3) - (duplicate_rows * 2)))

    return ImportAnalysis(
        dataset_type=dataset_type,
        source_format=source_format,
        records_received=len(rows),
        duplicate_rows=duplicate_rows,
        inferred_types=infer_column_types(rows),
        suggested_mapping=mapping,
        quality_score=quality_score,
        warnings=warnings,
        issues=issues[:50],
        sample_rows=rows[:5],
        unique_rows=list(deduplicated.values()) if deduplicated else rows,
    )


def analyze_import_file(
    file_bytes: bytes,
    *,
    filename: str,
    dataset_type: DatasetType,
    provided_mapping: dict[str, str | None] | None = None,
) -> ImportAnalysis:
    rows, source_format = parse_tabular_bytes(file_bytes, filename)
    mapping = detect_mapping(list(rows[0].keys()), dataset_type) if rows else {}
    if provided_mapping:
        mapping = {**mapping, **provided_mapping}
    return analyze_rows(rows, dataset_type=dataset_type, source_format=source_format, mapping=mapping)


def _persist_quality_issues(
    db: Session,
    batch: ImportBatch,
    issues: list[schemas.DataQualityIssueRecord],
) -> None:
    for issue in issues:
        db.add(
            DataQualityIssue(
                import_batch_id=batch.id,
                severity=issue.severity,
                issue_type=issue.issue_type,
                field_name=issue.field_name,
                row_number=issue.row_number,
                message=issue.message,
                sample_value=issue.sample_value,
            )
        )


def import_beneficiaries(
    db: Session,
    program: Program,
    rows: list[dict[str, str]],
    mapping: dict[str, str | None],
    filename: str,
    *,
    source_format: SourceFormat = "csv",
) -> ImportBatch:
    analysis = analyze_rows(rows, dataset_type="beneficiaries", source_format=source_format, mapping=mapping)
    batch = ImportBatch(
        program_id=program.id,
        dataset_type="beneficiaries",
        source_format=source_format,
        filename=filename,
        records_received=analysis.records_received,
        duplicates_detected=analysis.duplicate_rows,
        resolved_mapping=mapping,
        quality_summary={
            "quality_score": analysis.quality_score,
            "error_count": sum(1 for issue in analysis.issues if issue.severity == "error"),
            "warning_count": sum(1 for issue in analysis.issues if issue.severity == "warning"),
        },
    )
    db.add(batch)
    db.flush()
    _persist_quality_issues(db, batch, analysis.issues)

    existing_beneficiaries = {
        beneficiary.external_id: beneficiary
        for beneficiary in db.scalars(select(Beneficiary).where(Beneficiary.program_id == program.id)).all()
    }

    processed = 0
    failed = 0
    warnings = list(analysis.warnings)

    for row_number, row in enumerate(analysis.unique_rows, start=2):
        try:
            external_id = _read_value(row, mapping.get("external_id"))
            full_name = _read_value(row, mapping.get("full_name"))
            enrollment_date = _parse_date(_read_value(row, mapping.get("enrollment_date")))
            region = _read_value(row, mapping.get("region"))

            if not external_id or not full_name or not enrollment_date or not region:
                raise ValueError("Missing one or more required beneficiary fields.")

            beneficiary = existing_beneficiaries.get(external_id)
            if beneficiary is None:
                beneficiary = Beneficiary(
                    program_id=program.id,
                    external_id=external_id,
                    full_name=full_name,
                    region=region,
                    enrollment_date=enrollment_date,
                )
                db.add(beneficiary)
                existing_beneficiaries[external_id] = beneficiary

            beneficiary.full_name = full_name
            beneficiary.gender = _read_value(row, mapping.get("gender"))
            beneficiary.region = region
            beneficiary.cohort = _read_value(row, mapping.get("cohort"))
            beneficiary.phase = _read_value(row, mapping.get("phase"))
            beneficiary.household_type = _read_value(row, mapping.get("household_type"))
            beneficiary.delivery_modality = _read_value(row, mapping.get("delivery_modality")) or program.delivery_modality
            beneficiary.enrollment_date = enrollment_date
            beneficiary.dropout_date = _parse_date(_read_value(row, mapping.get("dropout_date")))
            beneficiary.completion_date = _parse_date(_read_value(row, mapping.get("completion_date")))
            beneficiary.status = (_read_value(row, mapping.get("status")) or beneficiary.status or "active").lower()
            beneficiary.household_size = _parse_int(_read_value(row, mapping.get("household_size")))
            beneficiary.pmt_score = _parse_float(_read_value(row, mapping.get("pmt_score")))
            beneficiary.food_insecurity_index = _parse_float(_read_value(row, mapping.get("food_insecurity_index")))
            beneficiary.distance_to_service_km = _parse_float(_read_value(row, mapping.get("distance_to_service_km")))
            beneficiary.preferred_contact_phone = _read_value(row, mapping.get("preferred_contact_phone"))
            beneficiary.preferred_contact_channel = _read_value(row, mapping.get("preferred_contact_channel"))
            beneficiary.assigned_case_worker = _read_value(row, mapping.get("assigned_case_worker"))
            beneficiary.assigned_site = _read_value(row, mapping.get("assigned_site"))
            beneficiary.current_note = _read_value(row, mapping.get("current_note"))
            beneficiary.household_stability_signal = _parse_int(_read_value(row, mapping.get("household_stability_signal")))
            beneficiary.economic_stress_signal = _parse_int(_read_value(row, mapping.get("economic_stress_signal")))
            beneficiary.family_support_signal = _parse_int(_read_value(row, mapping.get("family_support_signal")))
            beneficiary.health_change_signal = _parse_int(_read_value(row, mapping.get("health_change_signal")))
            beneficiary.motivation_signal = _parse_int(_read_value(row, mapping.get("motivation_signal")))
            opted_out = _parse_bool(_read_value(row, mapping.get("opted_out")))
            beneficiary.opted_out = opted_out if opted_out is not None else beneficiary.opted_out
            consent_status = _read_value(row, mapping.get("modeling_consent_status"))
            if consent_status:
                beneficiary.modeling_consent_status = consent_status.strip().lower()
            consent_method = _read_value(row, mapping.get("consent_method"))
            if consent_method:
                beneficiary.consent_method = consent_method
            consent_note = _read_value(row, mapping.get("consent_note"))
            if consent_note:
                beneficiary.consent_note = consent_note
            processed += 1
        except ValueError as exc:
            failed += 1
            warning_message = f"Row {row_number}: {exc}"
            warnings.append(warning_message)
            db.add(
                DataQualityIssue(
                    import_batch_id=batch.id,
                    severity="error",
                    issue_type="row_rejected",
                    row_number=row_number,
                    message=warning_message,
                )
            )

    batch.records_processed = processed
    batch.records_failed = failed
    batch.warning_preview = warnings[:10]
    db.commit()
    db.refresh(batch)
    return batch


def import_events(
    db: Session,
    program: Program,
    rows: list[dict[str, str]],
    mapping: dict[str, str | None],
    filename: str,
    *,
    source_format: SourceFormat = "csv",
) -> ImportBatch:
    analysis = analyze_rows(rows, dataset_type="events", source_format=source_format, mapping=mapping)
    batch = ImportBatch(
        program_id=program.id,
        dataset_type="events",
        source_format=source_format,
        filename=filename,
        records_received=analysis.records_received,
        duplicates_detected=analysis.duplicate_rows,
        resolved_mapping=mapping,
        quality_summary={
            "quality_score": analysis.quality_score,
            "error_count": sum(1 for issue in analysis.issues if issue.severity == "error"),
            "warning_count": sum(1 for issue in analysis.issues if issue.severity == "warning"),
        },
    )
    db.add(batch)
    db.flush()
    _persist_quality_issues(db, batch, analysis.issues)

    beneficiaries = {
        beneficiary.external_id: beneficiary
        for beneficiary in db.scalars(select(Beneficiary).where(Beneficiary.program_id == program.id)).all()
    }
    existing_events = {
        (event.beneficiary_id, event.event_date, event.event_type.lower()): event
        for event in db.scalars(
            select(MonitoringEvent)
            .join(Beneficiary, MonitoringEvent.beneficiary_id == Beneficiary.id)
            .where(Beneficiary.program_id == program.id)
        ).all()
    }

    processed = 0
    failed = 0
    warnings = list(analysis.warnings)

    for row_number, row in enumerate(analysis.unique_rows, start=2):
        try:
            external_id = _read_value(row, mapping.get("external_id"))
            event_date = _parse_date(_read_value(row, mapping.get("event_date")))
            event_type = _read_value(row, mapping.get("event_type"))

            if not external_id or not event_date or not event_type:
                raise ValueError("Missing one or more required event fields.")

            beneficiary = beneficiaries.get(external_id)
            if beneficiary is None:
                raise ValueError(f"Beneficiary {external_id} does not exist in this program.")

            success_value = _parse_bool(_read_value(row, mapping.get("successful")))
            response_value = _parse_bool(_read_value(row, mapping.get("response_received")))
            normalized_event_type = event_type.lower()
            event_key = (beneficiary.id, event_date, normalized_event_type)
            event = existing_events.get(event_key)

            if event is None:
                event = MonitoringEvent(
                    beneficiary_id=beneficiary.id,
                    event_date=event_date,
                    event_type=normalized_event_type,
                )
                db.add(event)
                existing_events[event_key] = event

            event.successful = True if success_value is None else success_value
            event.response_received = response_value
            event.source = _read_value(row, mapping.get("source"))
            event.notes = _read_value(row, mapping.get("notes"))
            processed += 1
        except ValueError as exc:
            failed += 1
            warning_message = f"Row {row_number}: {exc}"
            warnings.append(warning_message)
            db.add(
                DataQualityIssue(
                    import_batch_id=batch.id,
                    severity="error",
                    issue_type="row_rejected",
                    row_number=row_number,
                    message=warning_message,
                )
            )

    batch.records_processed = processed
    batch.records_failed = failed
    batch.warning_preview = warnings[:10]
    db.commit()
    db.refresh(batch)
    return batch


def list_quality_issues(db: Session, import_batch_id: str, limit: int = 50) -> list[schemas.DataQualityIssueRecord]:
    statement = (
        select(DataQualityIssue)
        .where(DataQualityIssue.import_batch_id == import_batch_id)
        .order_by(DataQualityIssue.created_at.desc())
        .limit(limit)
    )
    return [
        schemas.DataQualityIssueRecord(
            id=item.id,
            severity=item.severity,  # type: ignore[arg-type]
            issue_type=item.issue_type,
            field_name=item.field_name,
            row_number=item.row_number,
            message=item.message,
            sample_value=item.sample_value,
            created_at=item.created_at,
        )
        for item in db.scalars(statement).all()
    ]
